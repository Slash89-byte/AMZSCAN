"""
Catalog Parser - Intelligent parsing of wholesaler catalog files

Handles:
- Variable table start positions (headers not always on row 1)
- Multiple file formats (CSV, Excel)
- Data validation and cleaning
- Robust error handling
"""

import csv
import logging
import re
from typing import Dict, List, Optional, Tuple, Any
from dataclasses import dataclass
from pathlib import Path
from io import StringIO

logger = logging.getLogger(__name__)


@dataclass
class CatalogData:
    """Parsed catalog data with metadata"""
    headers: List[str]
    rows: List[Dict[str, str]]
    header_row_index: int
    total_rows: int
    file_format: str
    encoding: str
    detected_currency: Optional[str] = None
    detected_separator: Optional[str] = None


class CatalogParser:
    """
    Intelligent parser for wholesaler catalog files.
    
    Features:
    - Auto-detects header row position
    - Supports CSV and Excel formats
    - Handles multiple encodings
    - Validates data structure
    """
    
    # Common patterns for detecting header rows
    HEADER_KEYWORDS = [
        'gtin', 'ean', 'upc', 'asin', 'sku', 'product', 'item',
        'price', 'cost', 'wholesale', 'retail', 'msrp',
        'stock', 'quantity', 'inventory', 'available',
        'name', 'title', 'description', 'brand', 'manufacturer',
        'category', 'type', 'department'
    ]
    
    # Currency symbols and codes
    CURRENCY_PATTERNS = {
        '€': 'EUR',
        '$': 'USD',
        '£': 'GBP',
        '¥': 'JPY',
        'CHF': 'CHF',
        'EUR': 'EUR',
        'USD': 'USD',
        'GBP': 'GBP',
        'JPY': 'JPY'
    }
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def parse_file(self, filepath: str, max_rows: int = 10000) -> CatalogData:
        """
        Parse catalog file and extract structured data.
        
        Args:
            filepath: Path to catalog file
            max_rows: Maximum rows to process
            
        Returns:
            CatalogData object with parsed information
            
        Raises:
            ValueError: If file format not supported or parsing fails
        """
        file_path = Path(filepath)
        
        if not file_path.exists():
            raise ValueError(f"File not found: {filepath}")
        
        # Determine file format
        file_format = file_path.suffix.lower()
        
        if file_format == '.csv':
            return self._parse_csv(filepath, max_rows)
        elif file_format in ['.xlsx', '.xls']:
            return self._parse_excel(filepath, max_rows)
        else:
            raise ValueError(f"Unsupported file format: {file_format}")
    
    def _parse_csv(self, filepath: str, max_rows: int) -> CatalogData:
        """Parse CSV file with auto-detection of encoding and delimiter"""
        
        # Try different encodings
        encodings = ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252', 'iso-8859-1']
        
        for encoding in encodings:
            try:
                with open(filepath, 'r', encoding=encoding) as f:
                    # Read first few lines to detect delimiter and header
                    sample = f.read(8192)
                    f.seek(0)
                    
                    # Detect CSV delimiter
                    delimiter = self._detect_delimiter(sample)
                    
                    # Parse CSV
                    csv_reader = csv.reader(f, delimiter=delimiter)
                    all_rows = list(csv_reader)
                    
                    if not all_rows:
                        continue
                    
                    # Find header row
                    header_row_index, headers = self._find_header_row(all_rows)
                    
                    if header_row_index == -1:
                        self.logger.warning(f"No header row found with encoding {encoding}")
                        continue
                    
                    # Extract data rows
                    data_rows = all_rows[header_row_index + 1:]
                    
                    # Limit rows
                    if len(data_rows) > max_rows:
                        self.logger.warning(f"Limiting to {max_rows} rows (found {len(data_rows)})")
                        data_rows = data_rows[:max_rows]
                    
                    # Convert to list of dictionaries
                    parsed_rows = []
                    for row in data_rows:
                        if len(row) >= len(headers):
                            row_dict = {headers[i]: row[i].strip() if i < len(row) else '' 
                                       for i in range(len(headers))}
                            parsed_rows.append(row_dict)
                    
                    # Detect currency
                    detected_currency = self._detect_currency(parsed_rows, headers)
                    
                    self.logger.info(f"Successfully parsed CSV with {encoding} encoding")
                    self.logger.info(f"Found {len(parsed_rows)} data rows with {len(headers)} columns")
                    
                    return CatalogData(
                        headers=headers,
                        rows=parsed_rows,
                        header_row_index=header_row_index,
                        total_rows=len(parsed_rows),
                        file_format='csv',
                        encoding=encoding,
                        detected_currency=detected_currency,
                        detected_separator=delimiter
                    )
                    
            except UnicodeDecodeError:
                continue
            except Exception as e:
                self.logger.error(f"Error parsing CSV with {encoding}: {str(e)}")
                continue
        
        raise ValueError("Failed to parse CSV file with any supported encoding")
    
    def _parse_excel(self, filepath: str, max_rows: int) -> CatalogData:
        """Parse Excel file (.xlsx, .xls)"""
        try:
            import openpyxl
            
            workbook = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            sheet = workbook.active
            
            # Read all rows
            all_rows = []
            for row in sheet.iter_rows(values_only=True):
                # Convert row to list of strings
                row_data = [str(cell) if cell is not None else '' for cell in row]
                all_rows.append(row_data)
            
            if not all_rows:
                raise ValueError("Excel file is empty")
            
            # Find header row
            header_row_index, headers = self._find_header_row(all_rows)
            
            if header_row_index == -1:
                raise ValueError("Could not find header row in Excel file")
            
            # Extract data rows
            data_rows = all_rows[header_row_index + 1:]
            
            # Limit rows
            if len(data_rows) > max_rows:
                self.logger.warning(f"Limiting to {max_rows} rows (found {len(data_rows)})")
                data_rows = data_rows[:max_rows]
            
            # Convert to list of dictionaries
            parsed_rows = []
            for row in data_rows:
                if len(row) >= len(headers):
                    row_dict = {headers[i]: str(row[i]).strip() if i < len(row) else '' 
                               for i in range(len(headers))}
                    parsed_rows.append(row_dict)
            
            # Detect currency
            detected_currency = self._detect_currency(parsed_rows, headers)
            
            self.logger.info(f"Successfully parsed Excel file")
            self.logger.info(f"Found {len(parsed_rows)} data rows with {len(headers)} columns")
            
            return CatalogData(
                headers=headers,
                rows=parsed_rows,
                header_row_index=header_row_index,
                total_rows=len(parsed_rows),
                file_format='excel',
                encoding='utf-8',
                detected_currency=detected_currency
            )
            
        except ImportError:
            raise ValueError("openpyxl package required for Excel files. Install: pip install openpyxl")
        except Exception as e:
            raise ValueError(f"Error parsing Excel file: {str(e)}")
    
    def _detect_delimiter(self, sample: str) -> str:
        """Detect CSV delimiter from sample"""
        # Common delimiters in order of preference
        delimiters = [',', ';', '\t', '|']
        
        # Count occurrences of each delimiter
        delimiter_counts = {}
        for delimiter in delimiters:
            lines = sample.split('\n')[:5]  # Check first 5 lines
            counts = [line.count(delimiter) for line in lines if line.strip()]
            if counts and max(counts) > 0:
                # Use average count (should be consistent for real delimiter)
                delimiter_counts[delimiter] = sum(counts) / len(counts)
        
        if not delimiter_counts:
            return ','  # Default to comma
        
        # Return delimiter with highest average count
        return max(delimiter_counts, key=delimiter_counts.get)
    
    def _find_header_row(self, rows: List[List[str]]) -> Tuple[int, List[str]]:
        """
        Find the header row by looking for common header keywords.
        
        Returns:
            Tuple of (row_index, headers) or (-1, []) if not found
        """
        # Search first 20 rows for header
        search_limit = min(20, len(rows))
        
        best_score = 0
        best_index = -1
        best_headers = []
        
        for i in range(search_limit):
            row = rows[i]
            
            # Skip empty rows
            if not any(cell.strip() for cell in row):
                continue
            
            # Calculate header score
            score = self._score_header_row(row)
            
            if score > best_score:
                best_score = score
                best_index = i
                best_headers = [cell.strip() for cell in row]
        
        # Require minimum score to consider it a valid header
        if best_score < 2:
            return -1, []
        
        return best_index, best_headers
    
    def _score_header_row(self, row: List[str]) -> int:
        """
        Score a row based on how likely it is to be a header row.
        
        Looks for common header keywords and patterns.
        """
        score = 0
        
        for cell in row:
            cell_lower = cell.lower().strip()
            
            # Check for header keywords
            for keyword in self.HEADER_KEYWORDS:
                if keyword in cell_lower:
                    score += 2
                    break
            
            # Check for typical header patterns
            if cell_lower and not cell_lower.replace('.', '').replace(',', '').isdigit():
                score += 1
            
            # Penalize rows with mostly numbers (likely data rows)
            if cell.replace('.', '').replace(',', '').replace('-', '').isdigit():
                score -= 1
        
        return score
    
    def _detect_currency(self, rows: List[Dict[str, str]], headers: List[str]) -> Optional[str]:
        """
        Detect currency from price columns.
        
        Returns:
            Currency code (EUR, USD, GBP, etc.) or None
        """
        # Look for price-related columns
        price_columns = []
        for header in headers:
            header_lower = header.lower()
            if any(keyword in header_lower for keyword in ['price', 'cost', 'wholesale', 'retail']):
                price_columns.append(header)
        
        if not price_columns:
            return None
        
        # Check first 10 rows for currency symbols
        check_rows = rows[:10]
        
        for row in check_rows:
            for col in price_columns:
                value = row.get(col, '')
                
                # Check for currency symbols/codes
                for symbol, code in self.CURRENCY_PATTERNS.items():
                    if symbol in value:
                        return code
        
        return None
    
    def validate_data(self, catalog_data: CatalogData) -> Dict[str, Any]:
        """
        Validate parsed catalog data.
        
        Returns:
            Dictionary with validation results and statistics
        """
        validation = {
            'is_valid': True,
            'errors': [],
            'warnings': [],
            'stats': {
                'total_rows': catalog_data.total_rows,
                'columns': len(catalog_data.headers),
                'empty_rows': 0,
                'complete_rows': 0
            }
        }
        
        # Check minimum requirements
        if catalog_data.total_rows == 0:
            validation['is_valid'] = False
            validation['errors'].append("No data rows found")
            return validation
        
        if len(catalog_data.headers) == 0:
            validation['is_valid'] = False
            validation['errors'].append("No headers found")
            return validation
        
        # Check for empty rows
        for row in catalog_data.rows:
            values = [v for v in row.values() if v.strip()]
            if not values:
                validation['stats']['empty_rows'] += 1
            elif len(values) == len(catalog_data.headers):
                validation['stats']['complete_rows'] += 1
        
        # Warn if many empty rows
        empty_ratio = validation['stats']['empty_rows'] / catalog_data.total_rows
        if empty_ratio > 0.1:
            validation['warnings'].append(f"{empty_ratio:.1%} of rows are empty")
        
        return validation


def create_sample_catalog(filepath: str, format: str = 'csv'):
    """
    Create a sample catalog file for testing.
    
    Args:
        filepath: Output file path
        format: 'csv' or 'excel'
    """
    # Sample data with common structure
    data = [
        ["Wholesale Catalog - Beauty Products"],  # Title row
        ["Exported: 2025-10-04"],  # Metadata row
        [],  # Empty row
        ["EAN", "Product Name", "Brand", "Category", "Wholesale Price €", "Stock", "MOQ"],  # Headers
        ["3614227991341", "L'Oréal Paris Lash Paradise Mascara", "L'Oréal", "Cosmetics", "€8.50", "150", "12"],
        ["3600523351534", "Garnier Micellar Water 400ml", "Garnier", "Skincare", "€6.20", "200", "24"],
        ["3337875597388", "Lancôme La Vie Est Belle 50ml", "Lancôme", "Fragrance", "€42.00", "50", "6"],
        ["8411061976111", "Nivea Soft Cream 200ml", "Nivea", "Skincare", "€3.80", "500", "48"],
    ]
    
    if format == 'csv':
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerows(data)
    elif format == 'excel':
        try:
            import openpyxl
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            for row in data:
                sheet.append(row)
            workbook.save(filepath)
        except ImportError:
            raise ValueError("openpyxl required for Excel files")
    else:
        raise ValueError(f"Unsupported format: {format}")
