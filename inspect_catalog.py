"""Inspect the real wholesaler catalog"""
import openpyxl

wb = openpyxl.load_workbook('assets/Catalog_for_Seller_RD997Z-7Z9O6.xlsx', read_only=True, data_only=True)
ws = wb.active

print(f'Sheet name: {ws.title}')
print(f'Max row: {ws.max_row}')
print(f'Max column: {ws.max_column}')
print('\nFirst 20 rows:')
print('='*150)

for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    # Show row with limited cell content
    row_display = []
    for cell in row:
        if cell is None:
            row_display.append('None')
        else:
            cell_str = str(cell)[:50]  # Limit to 50 chars
            row_display.append(cell_str)
    print(f'Row {i}: {" | ".join(row_display)}')
    if i >= 20:
        break
